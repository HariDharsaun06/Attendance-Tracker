from flask import Flask, request, render_template
import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

app = Flask(__name__)

# Configuration
FILE_PATH = r"C:\Users\harid\OneDrive\Documents\satdb\Book1.xlsx"
STAFF_MAILS = ['haridharsaun1806@gmail.com', 'haridharsaun18personal@gmail.com', 'haridharsaun.rl2023cse@sece.ac.in']
WARNINGS = {
    1: "warning!!! you can take only one more day leave for C++ class",
    2: "warning!!! you can take only one more day leave for Python class",
    3: "warning!!! you can take only one more day leave for DS class"
}

def savefile(book):
    book.save(FILE_PATH)
    print("saved!")

def mailstu(li, msg):
    from_id = 'sat74637@gmail.com'
    pwd = 'fkjl kfko pfjo qenm'
    s = smtplib.SMTP('smtp.gmail.com', 587, timeout=120)
    s.starttls()
    s.login(from_id, pwd)

    for to_id in li:
        message = MIMEMultipart()
        message['Subject'] = 'Attendance report'
        message.attach(MIMEText(msg, 'plain'))
        content = message.as_string()
        s.sendmail(from_id, to_id, content)

    s.quit()
    print("mail sent to students")

def mailstaff(mail_id, msg):
    from_id = 'sat74637@gmail.com'
    pwd = 'fkjl kfko pfjo qenm'
    message = MIMEMultipart()
    message['Subject'] = 'Lack of attendance report'
    message.attach(MIMEText(msg, 'plain'))

    s = smtplib.SMTP('smtp.gmail.com', 587, timeout=120)
    s.starttls()
    s.login(from_id, pwd)
    content = message.as_string()
    s.sendmail(from_id, mail_id, content)
    s.quit()
    print('Mail Sent to staff')

def check(no_of_days, row_num, b, sheet):
    l1 = []
    l2 = ""
    l3 = []

    for student in range(len(row_num)):
        if no_of_days[student] == 2:
            l1.append(sheet.cell(row=row_num[student], column=2).value)
            mailstu(l1, WARNINGS[b])

        elif no_of_days[student] > 2:
            l2 += str(sheet.cell(row=row_num[student], column=1).value) + " "
            l3.append(sheet.cell(row=row_num[student], column=2).value)

    if l2 and l3:
        subject = "C++" if b == 1 else "Python" if b == 2 else "Data Structure"
        msg1 = f"you have lack of attendance in {subject}!!!"
        msg2 = f"the following students have lack of attendance in your subject: {l2}"

        mailstu(l3, msg1)
        staff_id = STAFF_MAILS[b - 1]
        mailstaff(staff_id, msg2)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    subject = int(request.form['subject'])
    no_of_absentees = int(request.form['no_of_absentees'])
    roll_nos = list(map(int, request.form['roll_nos'].split()))

    book = openpyxl.load_workbook(FILE_PATH)
    sheet = book['Sheet1']
    r = sheet.max_row
    row_num = []
    no_of_days = []

    for student in roll_nos:
        for i in range(2, r + 1):
            if sheet.cell(row=i, column=1).value == student:
                col = 3 if subject == 1 else 4 if subject == 2 else 5
                m = sheet.cell(row=i, column=col).value + 1
                sheet.cell(row=i, column=col).value = m
                savefile(book)
                no_of_days.append(m)
                row_num.append(i)

    check(no_of_days, row_num, subject, sheet)
    return "Attendance recorded and emails sent!"

if __name__ == '__main__':
    app.run(debug=True)
